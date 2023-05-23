
# 2.34.59
# напечататс писок страниц файла xl создат список кортежей(имя стрб кол-во заполненых ячеек и



import openpyxl
from openpyxl import Workbook
wb=Workbook()                         # создаем экземпляр класса Workbook
# wb.save('xl_for_2023_05_22_2.xlsx')   # создаем файл 'xl_for_2023_05_22_2.xlsx'

wb=openpyxl.load_workbook('xl_for_2023_05_22_2.xlsx')
s=wb.sheetnames
res=[]
for i in s:
    ws = wb[i]
    res.append((ws.title, ws.max_row*ws.max_column))
print(sorted(res))
print(sorted(res, key= lambda x: -x[1]))
