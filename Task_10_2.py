            # Task_10_1
''' 
Дан файл *.xlsx. В первом листе фамилии и выработка по дням за один период.
В втором листе тоже самое за второй период времени.
Фамилии могу повторяться в доном листе несколько раз,
а могут быть разными.
Необходимо создать третий лист, который суммирует выработку из первых двух.
Список должен быть отсортирован по фамилиям.
'''

import openpyxl
from openpyxl import Workbook

# Ввод исходных данных (составление исходного файла)

wb=Workbook()                  # создать экземпляр класса Workbook
wb.save('test_10_2.xlsx')      # создать файл *.xlsx

# .......заполнение первого листа

lst_2=[('Иванов',100,200,100), ('Петров',400,0,0), ('Иванов',200,100,100)]

wb=openpyxl.load_workbook('test_10_2.xlsx')  # открыть файл *.xlsx
ws = wb.active                               # cделать единственный лист активным
ws.title = "even People and results I"       # задать имя активному листу

i=1
for k in range(0, len(lst_2)):
    ws['A' + str(i)] = lst_2[k][0]
    ws['B' + str(i)] = lst_2[k][1]
    ws['C' + str(i)] = lst_2[k][2]
    ws['D' + str(i)] = lst_2[k][3]
    i=i+1
wb.save('test_10_2.xlsx')      # сохранить файл *.xlsx

# .......заполнение ВТОРОГО листа

lst_3=[('Иванов',0,0,0), ('Петров',400,0,0), ('Сидоров',400,200,200)]

wb=openpyxl.load_workbook('test_10_2.xlsx')    # открыть файл *.xlsx
wb.create_sheet("even People and results II")  # Создаётся лист с названием ...
ws=wb["even People and results II"]            # ????? Назначение активного листа с названием ...

i=1
for k in range(0, len(lst_3)):
    ws['A' + str(i)] = lst_3[k][0]
    ws['B' + str(i)] = lst_3[k][1]
    ws['C' + str(i)] = lst_3[k][2]
    ws['D' + str(i)] = lst_3[k][3]
    i=i+1
wb.save('test_10_2.xlsx')      # сохранить файл *.xlsx

# Чтение файла ..........................

ws=wb["even People and results I"]            # ????? Назначение активного листа с названием ...

min_column = ws.min_column
max_column = ws.max_column
min_row = ws.min_row
max_row = ws.max_row
d={}
for i in range(min_row, max_row+1):
    for j in range(min_column, max_column):
        cel_val=ws.cell(row=i, column=j+1).value
        cel_nam=ws.cell(row=i, column=1).value
        d[cel_nam]=d.get(cel_nam,0)+cel_val
ws=wb["even People and results II"]            # ????? Назначение активного листа с названием ...
for i in range(min_row, max_row+1):
    for j in range(min_column, max_column):
        cel_val=ws.cell(row=i, column=j+1).value
        cel_nam=ws.cell(row=i, column=1).value
        d[cel_nam]=d.get(cel_nam,0)+cel_val
print(d)

lst=[]
lst = sorted(d.items())
lst.append(('ИТОГО', sum(d.values())))
print(lst)

# Запись итогов по каждому сотруднику и за всех прграммистовв файл *.xlsx..........................
wb.create_sheet("sum")  # Создаётся лист с названием ...
ws=wb["sum"]            # ????? Назначение активного листа с названием ...

k=1
for i in range(0, len(lst)):
    ws['A' + str(k)].value = lst[i][0]
    ws['B' + str(k)].value = lst[i][1]
    k+=1
wb.save('test_10_2.xlsx')
