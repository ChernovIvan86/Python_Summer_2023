# словарь имеет три уровня вложенности. напечатайте все ключи иэлеементы
import functools

import openpyxl

dct={1:{11:{111:1111}},
   2:{22:222}, 3:{33:333},
   4:444}
print(dct)

for k, v in dct.items():
    print(k,v)
    if type(v) == dict:
        for p, q in v.items():
            print(p, q)
            if type(q) == dict:
                for x, y in q.items():
                    print(x, y)

# а если число уровней вложенности не известно (пользовать рекурсии)

# functools.cmp_to_key() - превращант функцию сравнения в функцию "key="

from functools import cmp_to_key
def xy(x, y):
    if x < y: return -1
    elif x == y: return 0
    else: return 1

lst=[100, -1, 0, 1000, 20, 5]
print(sorted(lst, key= cmp_to_key(xy)))

# дан список целых чисел больше 1. отсортируйте их по величине первого простого делителя
# т.е. всё что делитьсяна 2, потом на 3(но не на 2), потом на 5(но не на 2 и 3)
def pp(x,y):
    i=2
    while x % i:
        i+=1
    j=2
    while y % j:
        j+= 1
    if i < j: return -1
    elif i == j: return 0
    else: return 1

lst=[31, 5, 6, 7, 25, 11, 10, 125, 1024]
print(sorted(lst, key=cmp_to_key(pp)))

# reduce(), принимает два обязательных параметра - ф-ю и список.
# принимает стоящую первым аргументом ф-ю для двух начальных элементов списка.
# а затем использует полученнное значение вместе соследующим значением списка,
# так пока список не закончится. итоговое значение не выводится
#
from functools import reduce
print(reduce(lambda  x, y : x+y, [1,2,3,4,5], 100))

# печать ...

st1='0'+'\n'
st2='1'+'\n'
st3='2'+'\n'
st4='3'+'\n'
st5 ='4'+'\n'
st6 ='\n'

f_for_2023_05_22 = open("f_for_2023_05_22", 'w+', encoding='UTF-8')
f_for_2023_05_22.write(st1+st2+st3+st4+st5+st6)
f_for_2023_05_22.close()

f_for_2023_05_22 = open("f_for_2023_05_22", 'w', encoding='UTF-8')
for i in range(5):
    print(i,i*i, file=f_for_2023_05_22)
    print(111,i,i*i )
f_for_2023_05_22.close()
# !!!!!!!!!!!!!!!!!!


# ..... 1.37.08
#st1='QWERTYUIOP{}[]'+'\n'
#st2='asdfghjkl;'+'\n'
#st3='ZXCVvvvbnm,./'+'\n'
#st4='qwerfghjbnm,['+'\n'
#st5 ='polkjhvcwerfghnjm'+'\n'
#st6 ='\n'

#f_for_2023_05_22_1 = open("f_for_2023_05_22_1", 'w+', encoding='UTF-8')
#f_for_2023_05_22_1.write(st1+st2+st3+st4+st5+st6)
#f_for_2023_05_22_1.close()

#with open("f_for_2023_05_22_1", 'w', encoding='UTF-8') as fin:
#    s=fin.readlines()

#with open("f_for_2023_05_22_2", 'w', encoding='UTF-8') as fout:
#    for i in s:
#        x=''.join(sorted(i.strip().split(), key=str.lower()))+'\n'
#        print(x, file=fout)
#with open("f_for_2023_05_22_2", encoding='UTF-8') as fi:
#    for i in fi:
#        print(i.strip())


# модуль openpyxl

from openpyxl import Workbook
import openpyxl
wb=Workbook()                         # создаем экземпляр класса Workbook
wb.save('xl_for_2023_05_22_1.xlsx')   # создаем файл 'xl_for_2023_05_22_1.xlsx'
wb=openpyxl.load_workbook('xl_for_2023_05_22_1.xlsx')
print(wb.sheetnames)
ws=wb.active
print(ws.title)
wb.create_sheet('New')

wb.save('xl_for_2023_05_22_1.xlsx')
print(wb.sheetnames)

import openpyxl
wb=openpyxl.load_workbook('xl_for_2023_05_22_1.xlsx')
ws=wb.active
ws['A1'].value = 100
ws['B2'].value = 200
ws['C3'].value = ws['A1'].value + ws['B2'].value
print(ws['C3'])

