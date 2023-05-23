            # Task_10_3
# Дан файл *.xlsx содержащий список людей (люди не повторяются) и результаты их работы.
# сформируйте новый лист с следующими значениями:
    # минимальное значение;
    # максимальное значение;
    # среднее арифметическое;
    # медиана (срединное значение)
# Для списка с нечётным количеством членов Ме[1,3,5,7,9]=5.
# Для списка с чётным количеством членов Ме[1,3,5,7]=4 ( (3+5)/2=4 ).

import openpyxl
from openpyxl import Workbook

# Ввод исходных данных (составление исходного файла)

wb=Workbook()                         # создать экземпляр класса Workbook
wb.save('xl_for_Task_10_3.xlsx')      # создать файл *.xlsx

lst_2=[('A',22), ('B',12), ('C',32), ('D',42), ('E',10)]

wb=openpyxl.load_workbook('xl_for_Task_10_3.xlsx')  # открыть файл *.xlsx
ws = wb.active                                      # cделать единственный лист активным
ws.title = "even People and results"                # задать имя активному листу

i=2
for k in range(0, len(lst_2)):
    ws['B' + str(i)] = lst_2[k][0]
    ws['C' + str(i)] = lst_2[k][1]
    i=i+1
wb.save('xl_for_Task_10_3.xlsx')      # сохранить файл *.xlsx

# ..........................

min_column = ws.min_column
max_column = ws.max_column
min_row = ws.min_row
max_row = ws.max_row

dct={}
lst=[]

for i in range(2,max_row+1):
    cel_val=ws.cell(row=i, column=3).value
    cel_nam=ws.cell(row=i, column=2).value
    dct[cel_nam]=cel_val
    lst.append((cel_nam, cel_val))
print(lst, dct)
# ..........................

mi_lst=min(lst, key=(lambda x : x[1]))
ma_lst=max(lst, key=(lambda x : x[1]))
m_lst=0
me_lst=0
if len(lst)!=0:
    for i in range(0, len(lst)):
        m_lst=m_lst+lst[i][1]
    m_lst=m_lst/len(lst)
    print(m_lst)
else:pass

lst=sorted(lst, key = (lambda x : x[1] ))

print(lst, len(lst))

if len(lst)%2==0:
    x1=int(len(lst) / 2)-1
    x2 =int(len(lst) / 2)
    me_lst=(lst[x1][1]+lst[x2][1])*0.5
else:
    me_lst = lst[int((len(lst) / 2) + 0.5)-1][1]
print(me_lst)

# ..........................
wb.create_sheet("analysis")  # Создаётся лист с названием ...
ws=wb["analysis"]            # ????? Назначение активного листа с названием ...

ws['A1'].value = 'min'
ws['B1'].value = mi_lst[0]
ws['C1'].value = mi_lst[1]

ws['A2'].value = 'max'
ws['B2'].value = ma_lst[0]
ws['C2'].value = ma_lst[1]

ws['A3'].value = 'M[]'
#ws['B2'].value = ma_lst[0]
ws['C3'].value = m_lst

ws['A4'].value = 'Me'
ws['B4'].value = lst[int((len(lst) / 2) + 0.5)-1][0]
ws['C4'].value = me_lst

wb.save('xl_for_Task_10_3.xlsx')
# ..........................




