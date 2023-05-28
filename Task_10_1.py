            # Task_10_1
# Дан файл *.xlsx содержащий список людей (люди повторяются) и результаты их работы.
# сформируйте новый лист с сумарными итогами по каждому сотруднику и за всех прграммистов
# Например на первом листе:
    # Иванов 100
    # Петров 400
    # Иванов 200
# На втором листе:
    # Иванов 300
    # Петров 400
    # ИТОГО 700

import openpyxl
from openpyxl import Workbook

# Ввод исходных данных (составление исходного файла)

wb=Workbook()                  # создать экземпляр класса Workbook
wb.save('test_10_1.xlsx')      # создать файл *.xlsx

lst_2=[('Иванов',100), ('Петров',400), ('Иванов',200)]

wb=openpyxl.load_workbook('test_10_1.xlsx')  # открыть файл *.xlsx
ws = wb.active                               # cделать единственный лист активным
ws.title = "even People and results"         # задать имя активному листу

i=1
for k in range(0, len(lst_2)):
    ws['A' + str(i)] = lst_2[k][0]
    ws['B' + str(i)] = lst_2[k][1]
    i=i+1
wb.save('test_10_1.xlsx')      # сохранить файл *.xlsx

# Чтение файла ..........................
min_column = ws.min_column
max_column = ws.max_column
min_row = ws.min_row
max_row = ws.max_row

lst=[]
dct={}
for i in range(min_row, max_row+1):
    for j in range(min_column, max_column):
        cel_val=ws.cell(row=i, column=j+1).value
        cel_nam=ws.cell(row=i, column=j).value
        #dct[cel_nam]=cel_val
        lst.append((cel_nam, cel_val))
print(lst)

# Вычисление итогов по каждому сотруднику .......

for i in range(0, len(lst)):
    if lst[i][0] in dct.keys():
        dct[lst[i][0]] = dct[lst[i][0]] + lst[i][1]
    else:
        dct[lst[i][0]] = lst[i][1]
print(dct)
lst=[(k, v) for k, v in dct.items()]
print(lst)

# Вычисление итогов за всех прграммистов.......
su=0
for i in range(0, len(lst)):
    su=su+lst[i][1]
lst.append(('ИТОГО', su))
print(lst)

# Запись итогов по каждому сотруднику и за всех прграммистовв файл *.xlsx..........................
wb.create_sheet("sum")  # Создаётся лист с названием ...
ws=wb["sum"]            # ????? Назначение активного листа с названием ...

k=1
for i in range(0, len(lst)):
    ws['A' + str(k)].value = lst[i][0]
    ws['B' + str(k)].value = lst[i][1]
    k+=1
wb.save('test_10_1.xlsx')

