            # Task_11_2
# Дан файл с расширением .csv, содержащий в каждой строчке:
    # номер,фамилия,имя,компания,зарплата.
# создайте файл .xlsx и перенести эти данные,
# предворительно отсортировав их по компании, по фамилии и имени.
# В конце списка добавить строку: ИТОГО и суммарное значение всех зарплат.

import csv
import openpyxl
from openpyxl import Workbook

# запись

with open('test_11_2.csv', 'w', encoding='utf-8') as w_file:
    file_writer=csv.writer(w_file, delimiter=',', lineterminator='\r')
    file_writer.writerow(['номер','фамилия','имя','компания','зарплата'])
    file_writer.writerow(['1','Иванов','Степан','комп_1','300'])
    file_writer.writerow(['3','Петров','Антон','комп_2','200'])
    file_writer.writerow(['2','Сидоров','Иван','комп_4','250'])
    file_writer.writerow(['4','Пупкин','Егор','комп_3','350'])
    file_writer.writerow(['5', 'Сидоров', 'Антон', 'комп_3', '350'])
    file_writer.writerow(['6', 'Сидоров', 'Альберт', 'комп_3', '350'])

# Чтение
lst=[]
with open('test_11_2.csv', 'r', encoding='utf-8') as r_file:
    reader = csv.reader(r_file)
    for row in reader:
        print(row)
        lst.append(tuple(row))

lst.pop(0)
lst.sort(key=lambda x: (x[3], x[1], x[2]))
print(lst)

# вычисление суммы зарплат
sm=0
for i in range(0, len(lst)):
    sm=sm+float(lst[i][4])
print(sm)

# создайте файла .xlsx и перенос данных
wb=Workbook()                         # создать экземпляр класса Workbook
wb.save('test_11_2.xlsx')             # создать файл *.xlsx

wb=openpyxl.load_workbook('test_11_2.xlsx')  # открыть файл *.xlsx
ws = wb.active                               # cделать единственный лист активным
ws.title = "Task_11_2"                       # задать имя активному листу

sm=0
for k in range(0, len(lst)):
    ws['B' + str(k+1)] = lst[k][0]
    ws['C' + str(k + 1)] = lst[k][1]
    ws['D' + str(k + 1)] = lst[k][2]
    ws['E' + str(k + 1)] = lst[k][3]
    ws['F' + str(k + 1)] = lst[k][4]
    ws['B' + str(k + 2)] = 'ИТОГО'
    sm=sm+float(lst[k][4])
    ws['F' + str(k + 2)] = sm
wb.save('test_11_2.xlsx')      # сохранить файл *.xlsx

